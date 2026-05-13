"""Excel importer - imports field definitions from Excel files.

Provides:
- ExcelImporter: Main class for importing form templates and field layouts from Excel
- ExcelParseError: Exception for Excel parsing errors
- ImportMapping: Configuration for column-to-field mapping

Example:
    >>> from src.documents.constructor.excel_importer import ExcelImporter
    >>> importer = ExcelImporter()
    >>> template = importer.import_template(Path("template.xlsx"))
    >>> fields = importer.import_field_layout(Path("layout.xlsx"))
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from src.documents.constructor.field_builder import FormField
    from src.documents.constructor.form_constructor import FormTemplate

logger = logging.getLogger(__name__)

# Optional dependency check
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Workbook = Any  # type: ignore
    Cell = Any  # type: ignore
    Worksheet = Any  # type: ignore


class ExcelParseError(Exception):
    """Ошибка парсинга Excel файла.

    Attributes:
        message: Сообщение об ошибке.
        sheet: Имя листа (если известно).
        row: Номер строки (если известно).
        column: Имя/номер колонки (если известно).
    """

    def __init__(
        self,
        message: str,
        sheet: str | None = None,
        row: int | None = None,
        column: str | None = None,
    ):
        super().__init__(message)
        self.sheet = sheet
        self.row = row
        self.column = column

    def __str__(self) -> str:
        location = ""
        if self.sheet:
            location = f" в листе '{self.sheet}'"
        if self.row:
            location += f", строка {self.row}"
        if self.column:
            location += f", колонка {self.column}"
        return f"{self.args[0]}{location}"


@dataclass
class ImportMapping:
    """Конфигурация маппинга колонок Excel на поля.

    Attributes:
        field_id_column: Имя/индекс колонки для field_id.
        field_type_column: Имя/индекс колонки для типа поля.
        label_column: Имя/индекс колонки для метки.
        required_column: Имя/индекс колонки для обязательности.
        position_x_column: Имя/индекс колонки для X позиции.
        position_y_column: Имя/индекс колонки для Y позиции.
        width_column: Имя/индекс колонки для ширины.
        height_column: Имя/индекс колонки для высоты.
        validation_pattern_column: Имя/индекс колонки для паттерна.
        default_value_column: Имя/индекс колонки для значения по умолчанию.
        max_length_column: Имя/индекс колонки для макс. длины.
        options_column: Имя/индекс колонки для опций (через разделитель).
        options_separator: Разделитель для опций (по умолчанию ";")
        header_row: Номер строки с заголовками (по умолчанию 1).
        data_start_row: Номер первой строки с данными (по умолчанию 2).
    """

    field_id_column: str | int = "A"
    field_type_column: str | int = "B"
    label_column: str | int = "C"
    required_column: str | int = "D"
    position_x_column: str | int = "E"
    position_y_column: str | int = "F"
    width_column: str | int = "G"
    height_column: str | int = "H"
    validation_pattern_column: str | int | None = None
    default_value_column: str | int | None = None
    max_length_column: str | int | None = None
    options_column: str | int | None = None
    options_separator: str = ";"
    header_row: int = 1
    data_start_row: int = 2


class ExcelImporter:
    """Импортёр форм из Excel файлов.

    Импортирует определения полей и шаблоны форм
    из файлов Excel (.xlsx).

    Attributes:
        mapping: Конфигурация маппинга колонок.
        type_code: Код типа документа (для import_template).
        subtype: Код подтипа (для import_template).
        series: Серия документа (для import_template).

    Example:
        >>> importer = ExcelImporter()
        >>> # Импорт шаблона
        >>> template = importer.import_template(Path("template.xlsx"))
        >>> # Импорт только полей
        >>> fields = importer.import_field_layout(Path("fields.xlsx"))
    """

    def __init__(
        self,
        mapping: ImportMapping | None = None,
        type_code: str = "",
        subtype: str = "",
        series: str = "",
    ) -> None:
        """Инициализирует импортёр.

        Args:
            mapping: Конфигурация маппинга колонок.
            type_code: Код типа документа.
            subtype: Код подтипа.
            series: Серия документа.

        Raises:
            ImportError: Если openpyxl не установлен.
        """
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for Excel import. Install it with: pip install openpyxl"
            )

        self.mapping = mapping or ImportMapping()
        self.type_code = type_code
        self.subtype = subtype
        self.series = series

        logger.debug("ExcelImporter initialized")

    def import_template(self, excel_path: Path) -> "FormTemplate":
        """Импортирует шаблон формы из Excel файла.

        Ожидает файл с полями формы, включая type_code, subtype, series
        либо переданные через конструктор.

        Args:
            excel_path: Путь к Excel файлу (.xlsx).

        Returns:
            Импортированный шаблон.

        Raises:
            FileNotFoundError: Если файл не найден.
            ExcelParseError: Если файл невалиден.
        """
        from src.documents.constructor.form_constructor import FormTemplate

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        try:
            wb = load_workbook(excel_path, data_only=True)

            # Ищем лист с метаданными или используем первый
            meta_sheet = self._get_sheet_by_name(wb, ["Meta", "Metadata", "meta", "metadata"])
            fields_sheet = self._get_sheet_by_name(wb, ["Fields", "fields", "Поля"])

            # Загружаем метаданные если есть
            type_code = self.type_code
            subtype = self.subtype
            series = self.series
            metadata: dict[str, Any] = {}

            if meta_sheet:
                meta = self._parse_metadata(meta_sheet)
                type_code = meta.get("type_code", type_code)
                subtype = meta.get("subtype", subtype)
                series = meta.get("series", series)
                metadata = {
                    k: v for k, v in meta.items() if k not in ("type_code", "subtype", "series")
                }

            if not type_code or not subtype or not series:
                raise ExcelParseError(
                    "Missing required metadata: type_code, subtype, series. "
                    "Provide via file, constructor, or 'Meta' sheet"
                )

            # Загружаем поля
            if fields_sheet is None:
                fields_sheet = wb.active

            if fields_sheet is None:
                raise ExcelParseError("No worksheet found in Excel file")

            field_defaults = self._parse_field_defaults(fields_sheet)

            wb.close()

            template = FormTemplate(
                type_code=type_code,
                subtype=subtype,
                series=series,
                field_defaults=field_defaults,
                metadata=metadata,
            )

            logger.info(f"Imported template from {excel_path}")
            return template

        except Exception as e:
            if isinstance(e, ExcelParseError):
                raise
            raise ExcelParseError(f"Failed to parse Excel file: {e}")

    def import_field_layout(self, excel_path: Path) -> list["FormField"]:
        """Импортирует расположение и свойства полей из Excel.

        Импортирует полную конфигурацию полей с позициями
        и свойствами для создания формы.

        Args:
            excel_path: Путь к Excel файлу (.xlsx).

        Returns:
            Список импортированных полей.

        Raises:
            FileNotFoundError: Если файл не найден.
            ExcelParseError: Если файл невалиден.
        """
        from src.documents.constructor.field_builder import FieldBuilder
        from src.documents.types.type_schema import FieldType

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            if ws is None:
                raise ExcelParseError("No active worksheet in Excel file")

            fields: list[FormField] = []
            row = self.mapping.data_start_row

            while True:
                # Получаем значения из текущей строки
                field_id = self._get_cell_value(ws, self.mapping.field_id_column, row)

                # Если field_id пустой, считаем что данные закончились
                if not field_id:
                    break

                try:
                    # Собираем поле через Builder
                    builder = FieldBuilder()
                    builder.with_id(str(field_id))

                    # Тип поля
                    field_type_val = self._get_cell_value(ws, self.mapping.field_type_column, row)
                    if field_type_val:
                        try:
                            field_type = FieldType(str(field_type_val).lower())
                            builder.with_type(field_type)
                        except ValueError:
                            logger.warning(f"Unknown field type: {field_type_val}")

                    # Метка
                    label = self._get_cell_value(ws, self.mapping.label_column, row)
                    if label:
                        builder.with_label(str(label))

                    # Обязательность
                    required_val = self._get_cell_value(ws, self.mapping.required_column, row)
                    if required_val is not None:
                        required = self._parse_bool(str(required_val))
                        builder.with_required(required)

                    # Позиция
                    x = self._get_cell_int(ws, self.mapping.position_x_column, row)
                    y = self._get_cell_int(ws, self.mapping.position_y_column, row)
                    width = self._get_cell_int(ws, self.mapping.width_column, row)
                    height = self._get_cell_int(ws, self.mapping.height_column, row, default=1)

                    if x is not None and y is not None and width is not None:
                        builder.with_position(x=x, y=y, width=width, height=height)

                    # Опциональные поля
                    if self.mapping.validation_pattern_column:
                        pattern = self._get_cell_value(
                            ws, self.mapping.validation_pattern_column, row
                        )
                        if pattern:
                            builder.with_validation(pattern=str(pattern))

                    if self.mapping.default_value_column:
                        default = self._get_cell_value(ws, self.mapping.default_value_column, row)
                        if default is not None:
                            builder.with_default(default)

                    if self.mapping.max_length_column:
                        max_len = self._get_cell_int(ws, self.mapping.max_length_column, row)
                        if max_len is not None:
                            builder.with_max_length(max_len)

                    if self.mapping.options_column:
                        options_val = self._get_cell_value(ws, self.mapping.options_column, row)
                        if options_val:
                            options = str(options_val).split(self.mapping.options_separator)
                            builder.with_options(*[opt.strip() for opt in options])

                    # Собираем поле
                    field = builder.build()
                    fields.append(field)

                except Exception as e:
                    raise ExcelParseError(
                        f"Failed to parse field at row {row}: {e}",
                        sheet=ws.title,
                        row=row,
                    )

                row += 1

            wb.close()
            logger.info(f"Imported {len(fields)} fields from {excel_path}")
            return fields

        except Exception as e:
            if isinstance(e, ExcelParseError):
                raise
            raise ExcelParseError(f"Failed to parse Excel file: {e}")

    def export_field_layout(
        self,
        fields: list["FormField"],
        excel_path: Path,
        sheet_name: str = "Fields",
    ) -> None:
        """Экспортирует поля в Excel файл.

        Args:
            fields: Список полей для экспорта.
            excel_path: Путь для сохранения.
            sheet_name: Имя листа.
        """
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Failed to create worksheet")
        ws.title = sheet_name

        # Заголовки
        headers = [
            "field_id",
            "field_type",
            "label",
            "required",
            "pos_x",
            "pos_y",
            "width",
            "height",
            "validation_pattern",
            "default_value",
            "max_length",
            "options",
        ]
        ws.append(headers)

        # Данные
        for field in fields:
            row = [
                field.field_id,
                field.field_type.value if field.field_type else "",
                field.label,
                "Yes" if field.required else "No",
                field.position.x if field.position else "",
                field.position.y if field.position else "",
                field.position.width if field.position else "",
                field.position.height if field.position else 1,
                field.validation_pattern or "",
                str(field.default_value) if field.default_value is not None else "",
                field.max_length if field.max_length else "",
                self.mapping.options_separator.join(field.options) if field.options else "",
            ]
            ws.append(row)

        excel_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(excel_path)
        logger.info(f"Exported {len(fields)} fields to {excel_path}")

    def _get_sheet_by_name(self, wb: Workbook, names: list[str]) -> Worksheet | None:
        """Находит лист по одному из возможных имён."""
        for name in names:
            if name in wb.sheetnames:
                return wb[name]
        return None

    def _parse_metadata(self, ws: Worksheet) -> dict[str, Any]:
        """Парсит лист метаданных."""
        metadata: dict[str, Any] = {}

        for row in ws.iter_rows(min_row=1, values_only=True):
            if len(row) >= 2 and row[0]:
                key = str(row[0]).lower().strip()
                value = row[1]
                if key == "type_code" or key == "type":
                    metadata["type_code"] = str(value) if value else ""
                elif key == "subtype":
                    metadata["subtype"] = str(value) if value else ""
                elif key == "series":
                    metadata["series"] = str(value) if value else ""
                else:
                    metadata[key] = value

        return metadata

    def _parse_field_defaults(self, ws: Worksheet) -> dict[str, Any]:
        """Парсит значения полей по умолчанию."""
        defaults: dict[str, Any] = {}

        # Ожидаем формат: field_id | value
        for row in ws.iter_rows(min_row=self.mapping.data_start_row, values_only=True):
            if len(row) >= 2 and row[0] and row[1] is not None:
                field_id = str(row[0])
                value = row[1]
                defaults[field_id] = value

        return defaults

    def _get_cell_value(self, ws: Worksheet, column: str | int, row: int) -> Any:
        """Получает значение ячейки."""
        if isinstance(column, int):
            cell = ws.cell(row=row, column=column)
        else:
            cell = ws[f"{column}{row}"]

        value = cell.value
        if value is None:
            return None

        # Убираем лишние пробелы для строк
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None

        return value

    def _get_cell_int(
        self,
        ws: Worksheet,
        column: str | int,
        row: int,
        default: int | None = None,
    ) -> int | None:
        """Получает целочисленное значение ячейки."""
        value = self._get_cell_value(ws, column, row)
        if value is None:
            return default

        try:
            return int(float(value))
        except (ValueError, TypeError):
            return default

    def _parse_bool(self, value: str) -> bool:
        """Парсит булево значение из строки."""
        value = value.lower().strip()
        return value in ("true", "yes", "1", "да", "д", "y", "+")
