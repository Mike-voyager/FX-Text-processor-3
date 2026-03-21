"""Excel importer for document fields.

Provides:
- ExcelMappingType: Type of Excel data source
- ExcelFieldMapping: Mapping configuration for a field
- ExcelImporter: Imports data from Excel files
"""

from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any


class ExcelMappingType(str, Enum):
    """Тип источника данных в Excel."""

    COLUMN = "column"  # Столбец (вертикальный диапазон)
    ROW = "row"  # Строка (горизонтальный диапазон)
    RANGE = "range"  # Произвольный диапазон


@dataclass
class ExcelFieldMapping:
    """Описание маппинга поля документа на данные Excel.

    Attributes:
        field_name: Имя поля в схеме документа.
        source_type: Тип источника данных.
        sheet_name: Имя листа (None = активный лист).
        range_ref: Ссылка на диапазон (A1:B10, B:B, 3:3).
        skip_empty: Пропускать пустые ячейки.
        trim: Удалять пробелы по краям.
        dtype: Тип данных: "auto", "str", "int", "float", "date".
    """

    field_name: str
    source_type: ExcelMappingType
    sheet_name: str | None = None
    range_ref: str = ""
    skip_empty: bool = True
    trim: bool = True
    dtype: str = "auto"


class ExcelImporter:
    """Импорт данных из файлов Microsoft Excel.

    Поддерживает форматы .xlsx и .xls через openpyxl.
    """

    def __init__(self, file_path: Path) -> None:
        """Открывает файл Excel для чтения.

        Args:
            file_path: Путь к файлу Excel.

        Raises:
            FileNotFoundError: Если файл не найден.
            ValueError: Если формат не поддерживается.
        """
        self._file_path = file_path

        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        if file_path.suffix.lower() not in (".xlsx", ".xls"):
            raise ValueError(f"Unsupported Excel format: {file_path.suffix}")

        # Lazy import - openpyxl may not be installed
        self._workbook: Any = None
        self._openpyxl: Any = None

    def _get_openpyxl(self) -> Any:
        """Ленивая загрузка openpyxl."""
        if self._openpyxl is None:
            try:
                import openpyxl

                self._openpyxl = openpyxl
                self._workbook = openpyxl.load_workbook(
                    self._file_path, data_only=True
                )
            except ImportError:
                raise ImportError(
                    "openpyxl is required for Excel import. "
                    "Install with: pip install openpyxl"
                )
        return self._openpyxl

    def get_sheets(self) -> list[str]:
        """Возвращает список имён листов в файле.

        Returns:
            Список имён листов.
        """
        openpyxl = self._get_openpyxl()
        return list(self._workbook.sheetnames)

    def preview_range(
        self, sheet: str, range_ref: str, limit: int = 10
    ) -> list[Any]:
        """Предварительный просмотр данных из диапазона.

        Args:
            sheet: Имя листа.
            range_ref: Ссылка на диапазон (A1:D10, B:B, 3:3).
            limit: Максимальное количество строк.

        Returns:
            Список строк (для RANGE - список списков).
        """
        openpyxl = self._get_openpyxl()

        ws = self._workbook[sheet]

        # Parse range reference
        if ":" in range_ref:
            # Range like A1:D10
            start_ref, end_ref = range_ref.split(":")
            values: list[list[Any]] = []
            for row in ws[start_ref:end_ref]:  # type: ignore[misc]
                row_data = [cell.value for cell in row]
                values.append(row_data)
                if len(values) >= limit:
                    break
            return values
        elif range_ref.isalpha():
            # Column like B
            openpyxl = self._get_openpyxl()
            col = openpyxl.utils.column_index_from_string(range_ref)
            values = []
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    values.append(cell.value)
                    if len(values) >= limit:
                        break
            return values
        elif range_ref.isdigit():
            # Row like 3
            row_num = int(range_ref)
            values = []
            row = ws[row_num]
            for cell in row:
                values.append(cell.value)
            return values
        else:
            # Single cell
            cell = ws[range_ref]
            return [cell.value]

    def apply_mappings(
        self,
        mappings: list[ExcelFieldMapping],
        form_data: dict[str, Any],
    ) -> dict[str, Any]:
        """Применяет маппинги - извлекает данные из Excel.

        Args:
            mappings: Список маппингов.
            form_data: Существующие данные формы.

        Returns:
            Обновлённый словарь данных формы.
        """
        openpyxl = self._get_openpyxl()

        result = dict(form_data)

        for mapping in mappings:
            ws = self._workbook[mapping.sheet_name] if mapping.sheet_name else self._workbook.active

            # Get data based on mapping type
            if mapping.source_type == ExcelMappingType.COLUMN:
                data = self._read_column(ws, mapping.range_ref, mapping)
            elif mapping.source_type == ExcelMappingType.ROW:
                data = self._read_row(ws, mapping.range_ref, mapping)
            elif mapping.source_type == ExcelMappingType.RANGE:
                data = self._read_range(ws, mapping.range_ref, mapping)
            else:
                data = None

            result[mapping.field_name] = data

        return result

    def _read_column(
        self, worksheet: Any, range_ref: str, mapping: ExcelFieldMapping
    ) -> list[Any]:
        """Читает данные из столбца."""
        import re

        match = re.match(r"^([A-Z]+)(\d+)?:([A-Z]+)(\d+)?$", range_ref)
        if match:
            start_col, start_row, end_col, end_row = match.groups()
            start_row = int(start_row) if start_row else 1
            end_row = int(end_row) if end_row else worksheet.max_row
        else:
            # Just column letter
            col_letter = range_ref
            start_row = 1
            end_row = worksheet.max_row

        openpyxl = self._get_openpyxl()
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        values = []

        for row in range(start_row, end_row + 1):
            cell = worksheet.cell(row, col_idx)
            value = cell.value
            if mapping.trim and isinstance(value, str):
                value = value.strip()
            if not mapping.skip_empty or value is not None:
                value = self._convert_dtype(value, mapping.dtype)
                values.append(value)

        return values

    def _read_row(
        self, worksheet: Any, range_ref: str, mapping: ExcelFieldMapping
    ) -> list[Any]:
        """Читает данные из строки."""
        import re

        match = re.match(r"^(\d+):(\d+)$", range_ref)
        if match:
            start_row, end_row = match.groups()
            start_row = int(start_row)
            end_row = int(end_row)
        else:
            # Just row number
            row_num = int(range_ref)
            start_row = end_row = row_num

        values = []
        for col in range(1, worksheet.max_column + 1):
            for row in range(start_row, end_row + 1):
                cell = worksheet.cell(row, col)
                value = cell.value
                if mapping.trim and isinstance(value, str):
                    value = value.strip()
                value = self._convert_dtype(value, mapping.dtype)
                values.append(value)

        return values

    def _read_range(
        self, worksheet: Any, range_ref: str, mapping: ExcelFieldMapping
    ) -> list[list[Any]]:
        """Читает данные из диапазона."""
        openpyxl = self._get_openpyxl()

        values = []
        for row in worksheet[range_ref]:
            row_data = []
            for cell in row:
                value = cell.value
                if mapping.trim and isinstance(value, str):
                    value = value.strip()
                if not mapping.skip_empty or value is not None:
                    value = self._convert_dtype(value, mapping.dtype)
                    row_data.append(value)
            if row_data:
                values.append(row_data)

        return values

    def _convert_dtype(self, value: Any, dtype: str) -> Any:
        """Конвертирует значение в указанный тип."""
        if value is None:
            return None

        if dtype == "auto":
            return value

        try:
            if dtype == "str":
                return str(value)
            elif dtype == "int":
                return int(value)
            elif dtype == "float":
                return float(value)
            elif dtype == "date":
                if isinstance(value, str):
                    # Try to parse date
                    from datetime import datetime

                    return datetime.fromisoformat(value)
                return value
        except (ValueError, TypeError):
            return value

        return value

    def close(self) -> None:
        """Закрывает файл."""
        if self._workbook:
            self._workbook.close()

    def __enter__(self) -> "ExcelImporter":
        return self

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        self.close()